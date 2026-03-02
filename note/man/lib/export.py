import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, PatternFill
from io import BytesIO

def wb_style(ws) :

    cell_align = Alignment(
        horizontal = "center",
        vertical = "center"
    )

    style_border = Side(style = "thin", color = "A5A5A5")

    cell_border = Border(
        left = style_border,
        right = style_border,
        top = style_border,
        bottom = style_border
    )

    cell_fill = PatternFill(
        start_color = "E5E5E5",
        end_color = "E5E5E5",
        fill_type = "solid"
    )

    for column_cells in ws.columns :
        length = max(len(str(cell.value)) for cell in column_cells)
        adjusted_length = (length + 2) * 1.1
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_length

    for row in ws.iter_rows(min_row = 1, max_col = ws.max_column, max_row = ws.max_row):
        for cell in row :
            cell.alignment = cell_align
            cell.border = cell_border

    for cell in ws[1] :
        cell.fill = cell_fill

    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

def wb_export(df, sheet_name) :

    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine = "openpyxl") as writer:
        df.to_excel(writer, index = False, sheet_name = sheet_name)
        wb_style(writer.book[sheet_name])

    buffer.seek(0)

    return buffer
